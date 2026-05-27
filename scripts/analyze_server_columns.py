from __future__ import annotations

import csv
import json
import os
import re
import urllib.parse
import urllib.request
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
GEN_DIR = ROOT / "data" / "generated"
GEN_DIR.mkdir(parents=True, exist_ok=True)

ENV_PATH = Path(os.environ["ENV_PATH"])
XLSX_PATH = Path(os.environ["XLSX_PATH"])
JSON_PATH = Path(os.environ["JSON_PATH"])
CSV_PATH = Path(os.environ["CSV_PATH"]) if os.environ.get("CSV_PATH") else None


def read_module_key() -> str:
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        if line.strip().startswith("BID_MODULE_KEY="):
            return line.split("=", 1)[1].strip().strip("\"'")
    raise RuntimeError("BID_MODULE_KEY missing")


def fetch_server_notice_rows(module_key: str) -> list[dict]:
    params = {
        "moduleKey": module_key,
        "isDefault": "Y",
        "containCancel": "N",
        "onlyGong": "Y",
    }
    url = (
        "https://bidding2.kr/api2/module/dingpago/bidDataOrigin_get.php?"
        + urllib.parse.urlencode(params)
    )
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/124 Safari/537.36"
            ),
            "Accept": "application/json,text/html,*/*",
            "Referer": "https://bidding2.kr/",
        },
    )
    with urllib.request.urlopen(req, timeout=80) as resp:
        body = resp.read().decode("utf-8")
    rows = json.loads(body)
    if not isinstance(rows, list):
        raise RuntimeError(f"A1 response is not list: {type(rows)}")
    return rows


def ordered_columns(rows: list[dict]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                out.append(key)
    return out


EMPTY_STRS = {"", "—", "-", "null", "NULL", "None"}


def is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() in EMPTY_STRS:
        return True
    return False


def raw_type_name(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return type(value).__name__


def looks_int(text: str) -> bool:
    return bool(re.fullmatch(r"[+-]?\d+", text))


def looks_decimal(text: str) -> bool:
    return bool(re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text))


def looks_datetime(text: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", text))


def looks_date(text: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", text))


def looks_json(text: str) -> bool:
    stripped = text.strip()
    return (stripped.startswith("{") and stripped.endswith("}")) or (
        stripped.startswith("[") and stripped.endswith("]")
    )


MONEY_TERMS = [
    "금액",
    "가격",
    "원가",
    "A값",
    "시평액",
    "예가",
    "공사비",
    "추정금액",
    "기초금액",
    "부가가치세",
    "예상투찰",
    "제공자원가",
    "순공사",
    "투찰금액",
    "배정금액",
    "보전비",
    "보험료",
    "관리비",
    "시험비",
    "수수료",
    "부금비",
    "자재금액",
    "관급자재금액",
    "도급자설치",
    "관급자설치",
]
RATE_TERMS = ["율", "비율", "투찰율", "사정률", "계수"]
SCORE_TERMS = ["점수"]
AMOUNT_TERMS = ["금액", "가격", "원가", "시평액", "예가", "부가가치세", "투찰금액", "자재금액"]
FLAG_TERMS = ["여부", "자동수집", "내역입찰", "종목_모두보유", "공고상태", "참가신청", "발표전", "안함"]
ID_TERMS = ["번호", "코드"]

USER_COLUMN_DECISIONS = {
    "종목금액합계": {
        "action": "exclude",
        "memo": "서버에는 있지만 현재 업무에서 쓰지 않으므로 표준 컬럼 CSV에는 추가하지 않는다.",
    },
    "입력일": {
        "status": "사용자확정",
        "display_format": "yy-mm-dd",
        "memo": "시간은 표시하지 않고 날짜까지만 표시한다.",
    },
    "난이도계수": {
        "status": "수정필요",
        "display_format": "0.00",
        "memo": "소수점 둘째자리까지 표시한다.",
    },
    "상호진출여부": {
        "status": "사용자확정",
        "display_format": "#",
        "memo": "체크값으로 바꾸지 않고 0/1 숫자로 유지한다.",
    },
}


def has_any(text: str, terms: list[str]) -> bool:
    return any(term in text for term in terms)


def infer_column_profile(col: str, values: list, total_rows: int, seq: int) -> OrderedDict:
    raw_types = Counter(raw_type_name(v) for v in values)
    empty_count = sum(1 for v in values if is_empty(v))
    nonempty = [v for v in values if not is_empty(v)]
    nonempty_str = [str(v).strip() for v in nonempty]

    sample_values: list[str] = []
    sample_seen: set[str] = set()
    for value in nonempty_str:
        if value not in sample_seen:
            sample_seen.add(value)
            sample_values.append(value)
            if len(sample_values) >= 8:
                break

    unique_count = len(set(nonempty_str))
    type_scores = Counter()
    if nonempty_str:
        for value in nonempty_str:
            if looks_datetime(value):
                type_scores["datetime"] += 1
            if looks_date(value):
                type_scores["date"] += 1
            if looks_int(value):
                type_scores["integer_string"] += 1
            if looks_decimal(value):
                type_scores["decimal_string"] += 1
            if looks_json(value):
                type_scores["json_string"] += 1

    n = len(nonempty_str) or 1
    inferred = "text"
    reason = "기본 문자열/혼합값"
    display = ""

    col_is_money = has_any(col, MONEY_TERMS)
    col_is_rate = has_any(col, RATE_TERMS) and not has_any(col, AMOUNT_TERMS)
    col_is_score = has_any(col, SCORE_TERMS)
    col_is_flag = has_any(col, FLAG_TERMS)
    flag_values = {"0", "1", "true", "false", "True", "False", "Y", "N", "y", "n"}
    is_flag_value_set = bool(nonempty_str) and set(nonempty_str).issubset(flag_values)

    if nonempty_str and type_scores["json_string"] / n >= 0.8:
        inferred = "json_text"
        reason = "대부분 JSON 형태 문자열"
    elif nonempty_str and type_scores["datetime"] / n >= 0.8:
        inferred = "datetime"
        reason = "대부분 YYYY-MM-DD HH:mm:ss"
        display = "yyyy-mm-dd h:mm"
    elif nonempty_str and type_scores["date"] / n >= 0.8:
        inferred = "date"
        reason = "대부분 YYYY-MM-DD"
        display = "yyyy-mm-dd"
    elif is_flag_value_set and col_is_flag:
        inferred = "boolean_flag"
        reason = "0/1 또는 Y/N 상태값"
        display = "선택/체크"
    elif nonempty_str and type_scores["decimal_string"] / n >= 0.95:
        nums = []
        for value in nonempty_str:
            try:
                nums.append(float(value))
            except ValueError:
                pass
        if col_is_score:
            inferred = "score_decimal"
            reason = "점수 계열 숫자"
            display = "0.##"
        elif col_is_rate:
            inferred = "rate_decimal"
            reason = "비율/율 계열 숫자"
            display = "0.###"
        elif col_is_money:
            inferred = "money_integer" if all(abs(x - int(x)) < 1e-9 for x in nums) else "money_decimal"
            reason = "금액/비용 계열 숫자"
            display = "#,##0"
        elif type_scores["integer_string"] / n >= 0.95:
            inferred = "integer"
            reason = "대부분 정수 문자열"
            display = "0"
        else:
            inferred = "number_decimal"
            reason = "대부분 소수/숫자 문자열"
            display = "0.###"
    elif is_flag_value_set and col_is_flag:
        inferred = "boolean_flag"
        reason = "0/1 또는 Y/N 상태값"
        display = "선택/체크"
    elif unique_count <= 12 and len(nonempty_str) >= 20:
        inferred = "enum_text"
        reason = "고유값이 적은 선택형 문자열"

    if col in {"공고번호", "앞_공고번호", "gongNumKey", "idx"} or (
        has_any(col, ID_TERMS) and col != "공고번호_차수"
    ):
        inferred = "code_or_identifier"
        reason = "식별자/코드 계열은 문자열 보존"
        display = "@"

    most_common = Counter(nonempty_str).most_common(5)
    return OrderedDict(
        [
            ("순번", seq),
            ("항목", col),
            ("추정입력형식", inferred),
            ("권장표시형식", display),
            ("추정이유", reason),
            ("원본타입분포", ", ".join(f"{k}:{v}" for k, v in raw_types.items())),
            ("전체건수", total_rows),
            ("빈값수", empty_count),
            ("빈값비율", round(empty_count / total_rows, 4) if total_rows else 0),
            ("고유값수", unique_count),
            ("예시값", " | ".join(sample_values[:5])),
            ("상위값", " | ".join(f"{v}({c})" for v, c in most_common)),
        ]
    )


def load_xlsx_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in row):
            continue
        out.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
    return out


def load_csv_rows(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def normalize_display_format(value) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def display_format_status(col: str, inferred: str, csv_format: str) -> str:
    decision = USER_COLUMN_DECISIONS.get(col)
    if decision and decision.get("status") == "사용자확정":
        return "사용자확정"
    if decision and decision.get("status") == "수정필요":
        expected = normalize_display_format(decision.get("display_format"))
        return "일치" if normalize_display_format(csv_format) == expected else "수정필요"

    fmt = normalize_display_format(csv_format)
    if not fmt:
        return "미지정"
    if "입력값은" in fmt or "마스터데이터" in fmt or "일치하는값" in fmt:
        return "입력규칙메모"
    if inferred == "datetime":
        return "일치" if "yy" in fmt and "h:mm" in fmt else "확인필요"
    if inferred == "date":
        return "일치" if "yy" in fmt and "h" not in fmt else "확인필요"
    if inferred in {"money_integer", "integer"}:
        return "일치" if "#,##0" in fmt or fmt == "0" or ("#" in fmt and "," in fmt) or fmt == "#" else "확인필요"
    if inferred in {"rate_decimal"}:
        return "일치" if "%" in fmt else "확인필요"
    if inferred in {"money_decimal", "number_decimal", "score_decimal"}:
        return "일치" if "#" in fmt or "0" in fmt else "확인필요"
    if inferred == "code_or_identifier":
        return "일치" if fmt == "@" else "확인필요"
    return "확인필요"


def confirmed_csv_rows(csv_rows: list[dict], server_set: set[str], csv_only: set[str]) -> list[OrderedDict]:
    if not csv_rows:
        return []

    headers = list(csv_rows[0].keys())
    out_rows: list[OrderedDict] = []
    for row in csv_rows:
        item = str(row.get("항목") or "").strip()
        decision = USER_COLUMN_DECISIONS.get(item, {})
        out = OrderedDict()
        for header in headers:
            value = row.get(header, "")
            if header == "표시형식" and decision.get("display_format"):
                value = decision["display_format"]
            out[header] = value

        memo = decision.get("memo", "")
        if item in csv_only:
            memo = "서버 원본에는 없고 후처리/검증용 컬럼으로 유지한다."
        out["서버공고일치"] = "Y" if item in server_set else "N"
        out["확정메모"] = memo
        out_rows.append(out)
    return out_rows


def write_csv(path: Path, rows: list[OrderedDict | dict]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    rows = fetch_server_notice_rows(read_module_key())
    cols = ordered_columns(rows)

    profiles = [
        infer_column_profile(col, [row.get(col) for row in rows], len(rows), idx)
        for idx, col in enumerate(cols, start=1)
    ]

    xlsx_rows = load_xlsx_rows(XLSX_PATH)
    xlsx_map = {
        str(row.get("항목")).strip(): row
        for row in xlsx_rows
        if row.get("항목") is not None and str(row.get("항목")).strip()
    }

    json_rows = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    json_map = {
        str(row.get("항목")).strip(): row
        for row in json_rows
        if row.get("항목") is not None and str(row.get("항목")).strip()
    }

    csv_rows = load_csv_rows(CSV_PATH) if CSV_PATH else []
    csv_map = {
        str(row.get("항목")).strip(): row
        for row in csv_rows
        if row.get("항목") is not None and str(row.get("항목")).strip()
    }

    server_set = set(cols)
    xlsx_set = set(xlsx_map)
    json_set = set(json_map)
    csv_set = set(csv_map)

    compare_rows = []
    for profile in profiles:
        col = profile["항목"]
        xr = xlsx_map.get(col)
        jr = json_map.get(col)
        cr = csv_map.get(col)
        csv_display_status = display_format_status(
            col, profile["추정입력형식"], "" if not cr else str(cr.get("표시형식") or "")
        )
        user_decision = USER_COLUMN_DECISIONS.get(col, {})
        compare_rows.append(
            OrderedDict(
                [
                    ("순번", profile["순번"]),
                    ("항목", col),
                    ("서버_추정입력형식", profile["추정입력형식"]),
                    ("서버_권장표시형식", profile["권장표시형식"]),
                    ("서버_빈값비율", profile["빈값비율"]),
                    ("서버_예시값", profile["예시값"]),
                    ("엑셀_존재", "Y" if xr else "N"),
                    ("엑셀_빈값처리", "" if not xr else str(xr.get("빈값처리") or "")),
                    ("엑셀_빈값처리후", "" if not xr else str(xr.get("빈값처리후") or "")),
                    ("엑셀_표시형식", "" if not xr else str(xr.get("표시형식") or "")),
                    ("CSV_존재", "Y" if cr else "N"),
                    ("CSV_표시형식", "" if not cr else str(cr.get("표시형식") or "")),
                    ("CSV_표시형식검토", "" if not cr else csv_display_status),
                    ("CSV_처리방법", "" if not cr else str(cr.get("처리방법") or "")),
                    ("CSV_우선순위", "" if not cr else str(cr.get("우선순위") or "")),
                    ("CSV_참조방법", "" if not cr else str(cr.get("참조방법") or "")),
                    ("CSV_공고관리표시", "" if not cr else str(cr.get("공고관리 표시") or "")),
                    ("CSV_상세정보입력", "" if not cr else str(cr.get("상세정보입력") or "")),
                    ("사용자확정_표시형식", str(user_decision.get("display_format") or "")),
                    ("사용자확정_메모", str(user_decision.get("memo") or "")),
                    ("JSON_존재", "Y" if jr else "N"),
                    ("JSON_처리방법", "" if not jr else str(jr.get("처리방법") or "")),
                    (
                        "JSON_우선순위",
                        ""
                        if not jr
                        else json.dumps(jr.get("우선순위"), ensure_ascii=False)
                        if isinstance(jr.get("우선순위"), (list, dict))
                        else str(jr.get("우선순위") or ""),
                    ),
                    ("JSON_표시형식", "" if not jr else str(jr.get("표시형식") or "")),
                    ("비교상태_엑셀", "일치" if xr else "서버만 있음"),
                    ("비교상태_CSV", "일치" if cr else "서버만 있음"),
                    ("비교상태_JSON", "일치" if jr else "서버만 있음"),
                ]
            )
        )

    server_missing_in_xlsx = [c for c in cols if c not in xlsx_set]
    xlsx_only = sorted(xlsx_set - server_set)
    server_missing_in_csv = [c for c in cols if c not in csv_set]
    csv_only = sorted(csv_set - server_set)
    server_missing_in_json = [c for c in cols if c not in json_set]
    json_only = sorted(json_set - server_set)
    csv_display_review = [
        row
        for row in compare_rows
        if row["CSV_존재"] == "Y" and row["CSV_표시형식검토"] == "확인필요"
    ]
    csv_display_changes = [
        row
        for row in compare_rows
        if row["CSV_존재"] == "Y" and row["CSV_표시형식검토"] == "수정필요"
    ]
    csv_user_confirmed = [
        row
        for row in compare_rows
        if row["CSV_존재"] == "Y" and row["CSV_표시형식검토"] == "사용자확정"
    ]
    csv_input_rule_notes = [
        row
        for row in compare_rows
        if row["CSV_존재"] == "Y" and row["CSV_표시형식검토"] == "입력규칙메모"
    ]

    write_csv(GEN_DIR / "server_notice_column_profiles.csv", profiles)
    write_csv(GEN_DIR / "server_vs_api_column_format_compare.csv", compare_rows)
    write_csv(GEN_DIR / "server_notice_raw_wide.csv", [OrderedDict((col, row.get(col)) for col in cols) for row in rows])
    if CSV_PATH:
        write_csv(
            GEN_DIR / "api_column_format_confirmed_draft.csv",
            confirmed_csv_rows(csv_rows, server_set, set(csv_only)),
        )

    summary = OrderedDict(
        [
            ("generated_at", datetime.now().isoformat(timespec="seconds")),
            ("server_rows", len(rows)),
            ("server_columns_count", len(cols)),
            ("xlsx_file", str(XLSX_PATH)),
            ("xlsx_rows", len(xlsx_rows)),
            ("xlsx_columns_matched", len(server_set & xlsx_set)),
            ("server_columns_missing_in_xlsx_count", len(server_missing_in_xlsx)),
            ("xlsx_only_count", len(xlsx_only)),
            ("csv_file", str(CSV_PATH) if CSV_PATH else ""),
            ("csv_rows", len(csv_rows)),
            ("csv_columns_matched", len(server_set & csv_set)),
            ("server_columns_missing_in_csv_count", len(server_missing_in_csv)),
            ("csv_only_count", len(csv_only)),
            ("csv_display_review_count", len(csv_display_review)),
            ("csv_display_change_count", len(csv_display_changes)),
            ("csv_user_confirmed_count", len(csv_user_confirmed)),
            ("csv_input_rule_note_count", len(csv_input_rule_notes)),
            ("json_file", str(JSON_PATH)),
            ("json_rows", len(json_rows)),
            ("json_columns_matched", len(server_set & json_set)),
            ("server_columns_missing_in_json_count", len(server_missing_in_json)),
            ("json_only_count", len(json_only)),
            ("server_columns", cols),
            ("server_columns_missing_in_xlsx", server_missing_in_xlsx),
            ("xlsx_only", xlsx_only),
            ("server_columns_missing_in_csv", server_missing_in_csv),
            ("csv_only", csv_only),
            (
                "csv_display_review",
                [
                    {
                        "항목": row["항목"],
                        "서버_추정입력형식": row["서버_추정입력형식"],
                        "서버_권장표시형식": row["서버_권장표시형식"],
                        "CSV_표시형식": row["CSV_표시형식"],
                        "서버_예시값": row["서버_예시값"],
                    }
                    for row in csv_display_review
                ],
            ),
            (
                "csv_display_changes",
                [
                    {
                        "항목": row["항목"],
                        "서버_추정입력형식": row["서버_추정입력형식"],
                        "현재_CSV_표시형식": row["CSV_표시형식"],
                        "확정_표시형식": row["사용자확정_표시형식"],
                        "확정_메모": row["사용자확정_메모"],
                        "서버_예시값": row["서버_예시값"],
                    }
                    for row in csv_display_changes
                ],
            ),
            (
                "csv_user_confirmed",
                [
                    {
                        "항목": row["항목"],
                        "서버_추정입력형식": row["서버_추정입력형식"],
                        "CSV_표시형식": row["CSV_표시형식"],
                        "확정_메모": row["사용자확정_메모"],
                        "서버_예시값": row["서버_예시값"],
                    }
                    for row in csv_user_confirmed
                ],
            ),
            (
                "csv_input_rule_notes",
                [
                    {
                        "항목": row["항목"],
                        "서버_추정입력형식": row["서버_추정입력형식"],
                        "CSV_표시형식": row["CSV_표시형식"],
                        "서버_예시값": row["서버_예시값"],
                    }
                    for row in csv_input_rule_notes
                ],
            ),
            ("server_columns_missing_in_json", server_missing_in_json),
            ("json_only", json_only),
        ]
    )
    (GEN_DIR / "server_vs_api_column_format_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    profile_by_col = {profile["항목"]: profile for profile in profiles}
    format_counts = Counter(profile["추정입력형식"] for profile in profiles)

    lines: list[str] = []
    lines.append("# 서버공고 컬럼 형식 비교\n\n")
    lines.append(f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
    lines.append("## 요약\n\n")
    lines.append(f"- 서버공고 응답: {len(rows)}건, {len(cols)}컬럼\n")
    lines.append(f"- 비교 엑셀: `{XLSX_PATH.name}` {len(xlsx_rows)}행\n")
    lines.append(f"- 서버 컬럼 중 엑셀에 있는 항목: {len(server_set & xlsx_set)}개\n")
    lines.append(f"- 서버 컬럼 중 엑셀에 없는 항목: {len(server_missing_in_xlsx)}개\n")
    lines.append(f"- 엑셀에만 있고 서버에는 없는 항목: {len(xlsx_only)}개\n")
    if CSV_PATH:
        lines.append(f"- 비교 CSV: `{CSV_PATH.name}` {len(csv_rows)}행\n")
        lines.append(f"- 서버 컬럼 중 CSV에 있는 항목: {len(server_set & csv_set)}개\n")
        lines.append(f"- 서버 컬럼 중 CSV에 없는 항목: {len(server_missing_in_csv)}개\n")
        lines.append(f"- CSV에만 있고 서버에는 없는 항목: {len(csv_only)}개\n")
        lines.append(f"- CSV 표시형식 확인 필요: {len(csv_display_review)}개\n")
        lines.append(f"- CSV 표시형식 수정 반영: {len(csv_display_changes)}개\n")
        lines.append(f"- CSV 표시형식 사용자 확정 유지: {len(csv_user_confirmed)}개\n")
        lines.append(f"- CSV 입력규칙 메모 항목: {len(csv_input_rule_notes)}개\n")
    lines.append(f"- JSON 마스터: `{JSON_PATH.name}` {len(json_rows)}행\n")
    lines.append(f"- 서버 컬럼 중 JSON 마스터에 있는 항목: {len(server_set & json_set)}개\n")
    lines.append(f"- 서버 컬럼 중 JSON 마스터에 없는 항목: {len(server_missing_in_json)}개\n")
    lines.append(f"- JSON 마스터에만 있고 서버에는 없는 항목: {len(json_only)}개\n\n")

    lines.append("## 서버공고 추정 입력형식 분포\n\n")
    for key, count in format_counts.most_common():
        lines.append(f"- `{key}`: {count}\n")

    if CSV_PATH:
        lines.append("\n## 사용자 확정 사항\n\n")
        lines.append("- `종목금액합계`: 서버에는 있지만 현재 쓰지 않으므로 표준 CSV에 추가하지 않는다.\n")
        lines.append("- `입력일`: 시간은 버리고 날짜까지만 표시한다. CSV `yy-mm-dd`를 유지한다.\n")
        lines.append("- `난이도계수`: 소수점 둘째자리까지 표시한다. 확정초안에는 `0.00`으로 반영한다.\n")
        lines.append("- `상호진출여부`: 체크값이 아니라 `0/1` 숫자로 유지한다. CSV `#`를 유지한다.\n")
        lines.append("- CSV에만 있는 12개 컬럼은 서버 원본 컬럼이 아니라 후처리/검증용 컬럼으로 유지한다.\n")

    lines.append("\n## 엑셀에 없는 서버공고 항목\n\n")
    for col in server_missing_in_xlsx:
        profile = profile_by_col[col]
        lines.append(f"- `{col}`: {profile['추정입력형식']}, 예시 `{profile['예시값']}`\n")

    lines.append("\n## 엑셀에만 있는 항목\n\n")
    for col in xlsx_only:
        lines.append(f"- `{col}`\n")

    if CSV_PATH:
        lines.append("\n## CSV에 없는 서버공고 항목\n\n")
        for col in server_missing_in_csv:
            profile = profile_by_col[col]
            lines.append(f"- `{col}`: {profile['추정입력형식']}, 예시 `{profile['예시값']}`\n")

        lines.append("\n## CSV에만 있는 항목\n\n")
        for col in csv_only:
            lines.append(f"- `{col}`\n")

        lines.append("\n## CSV 표시형식 확인 필요 항목\n\n")
        if not csv_display_review:
            lines.append("- 없음\n")
        else:
            for row in csv_display_review:
                lines.append(
                    f"- `{row['항목']}`: 서버 `{row['서버_추정입력형식']}`, "
                    f"권장 `{row['서버_권장표시형식']}`, CSV `{row['CSV_표시형식']}`, "
                    f"예시 `{row['서버_예시값']}`\n"
                )

        lines.append("\n## CSV 표시형식 수정 반영 항목\n\n")
        for row in csv_display_changes:
            lines.append(
                f"- `{row['항목']}`: CSV `{row['CSV_표시형식']}` -> 확정 `{row['사용자확정_표시형식']}`, "
                f"이유 `{row['사용자확정_메모']}`, 예시 `{row['서버_예시값']}`\n"
            )

        lines.append("\n## CSV 표시형식 사용자 확정 유지 항목\n\n")
        for row in csv_user_confirmed:
            lines.append(
                f"- `{row['항목']}`: CSV `{row['CSV_표시형식']}` 유지, "
                f"이유 `{row['사용자확정_메모']}`, 예시 `{row['서버_예시값']}`\n"
            )

        lines.append("\n## CSV 입력규칙 메모 항목\n\n")
        for row in csv_input_rule_notes:
            lines.append(
                f"- `{row['항목']}`: CSV `{row['CSV_표시형식']}`, "
                f"예시 `{row['서버_예시값']}`\n"
            )

    lines.append("\n## JSON 마스터에 없는 서버공고 항목\n\n")
    for col in server_missing_in_json:
        profile = profile_by_col[col]
        lines.append(f"- `{col}`: {profile['추정입력형식']}, 예시 `{profile['예시값']}`\n")

    lines.append("\n## 산출 파일\n\n")
    lines.append("- `data/generated/server_notice_column_profiles.csv`: 서버공고 118컬럼별 값 형식 프로파일\n")
    lines.append("- `data/generated/server_vs_api_column_format_compare.csv`: 서버공고 vs 엑셀/JSON 비교표\n")
    lines.append("- `data/generated/server_vs_api_column_format_summary.json`: 비교 요약 JSON\n")
    if CSV_PATH:
        lines.append("- `data/generated/api_column_format_confirmed_draft.csv`: 사용자 답변을 반영한 CSV 확정초안\n")
    (ROOT / "docs" / "11_서버공고_컬럼형식_비교.md").write_text(
        "".join(lines), encoding="utf-8"
    )

    print(
        json.dumps(
            {
                "server_rows": len(rows),
                "server_columns": len(cols),
                "xlsx_rows": len(xlsx_rows),
                "xlsx_matched": len(server_set & xlsx_set),
                "missing_in_xlsx": len(server_missing_in_xlsx),
                "xlsx_only": len(xlsx_only),
                "csv_rows": len(csv_rows),
                "csv_matched": len(server_set & csv_set),
                "missing_in_csv": len(server_missing_in_csv),
                "csv_only": len(csv_only),
                "csv_display_review": len(csv_display_review),
                "csv_display_changes": len(csv_display_changes),
                "csv_user_confirmed": len(csv_user_confirmed),
                "csv_input_rule_notes": len(csv_input_rule_notes),
                "json_rows": len(json_rows),
                "json_matched": len(server_set & json_set),
                "missing_in_json": len(server_missing_in_json),
                "json_only": len(json_only),
                "format_counts": format_counts,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
